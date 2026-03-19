"""XLSX → PDF 변환 유틸리티 (openpyxl + reportlab)

원본 XLSX의 테두리, 병합 셀, 정렬, 폰트를 읽어서 PDF로 재현한다.
테두리가 없는 셀(제목 영역 등)에는 눈금선을 그리지 않는다.
"""

import os
import glob
from pathlib import Path


def xlsx_to_pdf(xlsx_path, output_pdf_path):
    """
    XLSX 파일을 PDF로 변환 — 원본 테두리/병합/정렬/폰트 반영

    Args:
        xlsx_path: 입력 XLSX 파일 경로
        output_pdf_path: 출력 PDF 파일 경로
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # 한글 폰트 등록
    fonts = _register_korean_fonts(pdfmetrics, TTFont)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # ── 페이지 설정 (방향, 여백, 스케일) ──
    orientation = getattr(ws.page_setup, 'orientation', None)
    is_landscape = (orientation == 'landscape')
    pagesize = landscape(A4) if is_landscape else A4

    margins = ws.page_margins
    left_margin = (margins.left or 0.7) * inch
    right_margin = (margins.right or 0.7) * inch
    top_margin = (margins.top or 0.75) * inch
    bottom_margin = (margins.bottom or 0.75) * inch

    scale_pct = getattr(ws.page_setup, 'scale', None) or 100
    scale = scale_pct / 100.0  # e.g. 90 → 0.9

    # ── 병합 셀 맵 구축 ──
    merge_map = {}       # merge_map[row][col] = 'origin' | 'hidden'
    merge_origins = {}   # (min_row, min_col) → (max_row, max_col)
    for merged_range in ws.merged_cells.ranges:
        min_r, min_c = merged_range.min_row, merged_range.min_col
        max_r, max_c = merged_range.max_row, merged_range.max_col
        merge_origins[(min_r, min_c)] = (max_r, max_c)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r not in merge_map:
                    merge_map[r] = {}
                if r == min_r and c == min_c:
                    merge_map[r][c] = 'origin'
                else:
                    merge_map[r][c] = 'hidden'

    # ── 열 너비 계산 ──
    available_width = pagesize[0] - left_margin - right_margin
    col_widths = _calc_col_widths(ws, max_col, available_width)

    # ── 셀 데이터 추출 (순수 문자열) + 개별 스타일 명령 ──
    table_data = []
    style_commands = []

    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            r = row_idx - 1  # reportlab 0-based row
            c = col_idx - 1  # reportlab 0-based col

            # 병합된 숨김 셀은 빈 문자열
            if merge_map.get(row_idx, {}).get(col_idx) == 'hidden':
                row_data.append('')
                continue

            # 셀 값
            value = cell.value
            if value is None:
                value = ''
            else:
                value = str(value)
                if ' 00:00:00' in value:
                    value = value.replace(' 00:00:00', '')

            row_data.append(value)

            # ── 폰트 → per-cell FONTNAME, FONTSIZE ──
            font = cell.font
            font_size = (font.size or 10) * scale
            font_bold = font.bold or False
            font_name = _resolve_font_name(font.name, font_bold, fonts)

            style_commands.append(('FONTNAME', (c, r), (c, r), font_name))
            style_commands.append(('FONTSIZE', (c, r), (c, r), font_size))
            style_commands.append(('LEADING', (c, r), (c, r), font_size * 1.2))

            # ── 정렬 ──
            align = 'LEFT'
            if cell.alignment and cell.alignment.horizontal:
                h = cell.alignment.horizontal.lower()
                if h == 'center':
                    align = 'CENTER'
                elif h == 'right':
                    align = 'RIGHT'
            style_commands.append(('ALIGN', (c, r), (c, r), align))
            style_commands.append(('VALIGN', (c, r), (c, r), 'MIDDLE'))

            # ── 테두리 ──
            border = cell.border
            if border.top.style and border.top.style != 'none':
                lw = _border_width(border.top.style)
                style_commands.append(
                    ('LINEABOVE', (c, r), (c, r), lw, colors.black))
            if border.bottom.style and border.bottom.style != 'none':
                lw = _border_width(border.bottom.style)
                style_commands.append(
                    ('LINEBELOW', (c, r), (c, r), lw, colors.black))
            if border.left.style and border.left.style != 'none':
                lw = _border_width(border.left.style)
                style_commands.append(
                    ('LINEBEFORE', (c, r), (c, r), lw, colors.black))
            if border.right.style and border.right.style != 'none':
                lw = _border_width(border.right.style)
                style_commands.append(
                    ('LINEAFTER', (c, r), (c, r), lw, colors.black))

        table_data.append(row_data)

    # 병합된 숨김 셀의 테두리도 읽기 (외곽선 정확도)
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            if merge_map.get(row_idx, {}).get(col_idx) == 'hidden':
                cell = ws.cell(row=row_idx, column=col_idx)
                r, c = row_idx - 1, col_idx - 1
                border = cell.border
                if border.top.style and border.top.style != 'none':
                    style_commands.append(
                        ('LINEABOVE', (c, r), (c, r),
                         _border_width(border.top.style), colors.black))
                if border.bottom.style and border.bottom.style != 'none':
                    style_commands.append(
                        ('LINEBELOW', (c, r), (c, r),
                         _border_width(border.bottom.style), colors.black))
                if border.left.style and border.left.style != 'none':
                    style_commands.append(
                        ('LINEBEFORE', (c, r), (c, r),
                         _border_width(border.left.style), colors.black))
                if border.right.style and border.right.style != 'none':
                    style_commands.append(
                        ('LINEAFTER', (c, r), (c, r),
                         _border_width(border.right.style), colors.black))

    # ── 병합 셀 → SPAN 명령 ──
    for (min_r, min_c), (max_r, max_c) in merge_origins.items():
        r0, c0 = min_r - 1, min_c - 1
        r1, c1 = max_r - 1, max_c - 1
        style_commands.append(('SPAN', (c0, r0), (c1, r1)))

    # ── 공통 패딩 (스케일 적용) ──
    pad = max(1, int(3 * scale))
    style_commands.extend([
        ('TOPPADDING', (0, 0), (-1, -1), pad),
        ('BOTTOMPADDING', (0, 0), (-1, -1), pad),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])

    # ── 행 높이 (스케일 적용, 페이지 맞춤) ──
    default_row_height = 18.0
    row_heights = []
    for row_idx in range(1, max_row + 1):
        dim = ws.row_dimensions.get(row_idx)
        h = dim.height if dim and dim.height else default_row_height
        row_heights.append(h * scale)

    # SimpleDocTemplate 내부 오버헤드(~12pt) 감안하여 여유 확보
    available_height = (pagesize[1] - top_margin - bottom_margin) * 0.97
    total_height = sum(row_heights)
    if total_height > available_height:
        shrink = available_height / total_height
        row_heights = [h * shrink for h in row_heights]

    # ── PDF 생성 ──
    doc = SimpleDocTemplate(
        str(output_pdf_path),
        pagesize=pagesize,
        leftMargin=left_margin, rightMargin=right_margin,
        topMargin=top_margin, bottomMargin=bottom_margin,
    )

    table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
    table.setStyle(TableStyle(style_commands))

    doc.build([table])


def _border_width(style_name):
    """Excel 테두리 스타일 → reportlab 선 두께"""
    return {
        'thin': 0.5,
        'medium': 1.0,
        'thick': 1.5,
        'hair': 0.25,
        'dotted': 0.5,
        'dashed': 0.5,
        'dashDot': 0.5,
        'dashDotDot': 0.5,
        'double': 1.0,
        'mediumDashed': 1.0,
        'mediumDashDot': 1.0,
        'mediumDashDotDot': 1.0,
        'slantDashDot': 1.0,
    }.get(style_name, 0.5)


def _calc_col_widths(ws, max_col, available_width):
    """Excel 열 너비를 비례 변환하여 PDF 열 너비 리스트 반환"""
    from openpyxl.utils import get_column_letter

    raw_widths = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width and dim.width > 0:
            raw_widths.append(dim.width)
        else:
            raw_widths.append(8.43)  # Excel 기본 열 너비

    total_raw = sum(raw_widths)
    if total_raw <= 0:
        return [available_width / max_col] * max_col

    # 비례 스케일링
    scaled = [w / total_raw * available_width for w in raw_widths]

    # 최소 열 너비 보장 (30pt ≈ 한글 3자)
    min_width = 30.0
    deficit = 0
    surplus_cols = []
    for i, w in enumerate(scaled):
        if w < min_width:
            deficit += min_width - w
            scaled[i] = min_width
        else:
            surplus_cols.append(i)

    if deficit > 0 and surplus_cols:
        per_col = deficit / len(surplus_cols)
        for i in surplus_cols:
            scaled[i] = max(min_width, scaled[i] - per_col)

    return scaled


def _register_korean_fonts(pdfmetrics, TTFont):
    """한글 폰트 등록 — KB금융체 Light/Medium/Bold + Display Bold"""
    import pathlib
    home = str(pathlib.Path.home())

    fonts = {}

    font_map = {
        'light': f"{home}/.local/share/fonts/kb-finance/KBfgTextL.ttf",
        'medium': f"{home}/.local/share/fonts/kb-finance/KBfgTextM.ttf",
        'bold': f"{home}/.local/share/fonts/kb-finance/KBfgTextB.ttf",
        'display_bold': f"{home}/.local/share/fonts/kb-finance/KBfgDisplayB.ttf",
    }

    reg_names = {
        'light': 'KBFontL',
        'medium': 'KBFontM',
        'bold': 'KBFontB',
        'display_bold': 'KBDisplayB',
    }

    for key, path in font_map.items():
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(reg_names[key], path))
                fonts[key] = reg_names[key]
            except Exception:
                pass

    # Fallback: NotoSansCJK
    if not fonts:
        noto_paths = [
            "/usr/share/fonts/google-noto-sans-cjk-ttc/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/google-noto/NotoSansCJK-Regular.ttc",
        ]
        for fp in noto_paths:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("NotoSansCJK", fp, subfontIndex=0))
                    for k in ['light', 'medium', 'bold', 'display_bold']:
                        fonts[k] = "NotoSansCJK"
                    break
                except Exception:
                    continue

        noto_fonts = glob.glob("/usr/share/fonts/**/NotoSans*CJK*.tt[cf]", recursive=True)
        if not fonts and noto_fonts:
            try:
                pdfmetrics.registerFont(TTFont("NotoSansCJK", noto_fonts[0], subfontIndex=0))
                for k in ['light', 'medium', 'bold', 'display_bold']:
                    fonts[k] = "NotoSansCJK"
            except Exception:
                pass

    if not fonts:
        fonts = {k: "Helvetica" for k in ['light', 'medium', 'bold', 'display_bold']}
        fonts['display_bold'] = "Helvetica-Bold"

    # 빠진 키 보완
    fallback = fonts.get('medium', fonts.get('light', 'Helvetica'))
    for key in ['light', 'medium', 'bold', 'display_bold']:
        if key not in fonts:
            fonts[key] = fallback

    return fonts


def _resolve_font_name(excel_font_name, is_bold, fonts):
    """Excel 폰트명 → reportlab 등록 폰트명 매핑"""
    if not excel_font_name:
        return fonts['medium'] if not is_bold else fonts['bold']

    name_lower = excel_font_name.lower()

    # KB금융 제목체
    if '제목' in excel_font_name or 'display' in name_lower:
        return fonts['display_bold']

    # KB금융 본문체
    if 'kb' in name_lower or '금융' in excel_font_name or '본문' in excel_font_name:
        if 'bold' in name_lower or is_bold:
            return fonts['bold']
        elif 'medium' in name_lower:
            return fonts['medium']
        else:
            return fonts['light']

    # 기타 폰트
    if is_bold:
        return fonts['bold']
    return fonts['light']
