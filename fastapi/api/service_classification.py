# -*- coding: utf-8 -*-
"""
    service_classification.py
    ~~~~~~~~~~~~~~~~~~~~~~~~~

    분류(장) 관리 API

    :copyright: (c) 2024 by wizice.
    :license:  wizice.com
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Dict, Any
import logging
import io
from datetime import datetime
from pydantic import BaseModel, field_validator
from .auth_middleware import get_current_user
from .timescaledb_manager_v2 import DatabaseConnectionManager
from settings import settings
from app_logger import get_logger

logger = get_logger(__name__)

# Pydantic 모델
class CategoryCreate(BaseModel):
    chapter_number: int
    name: str

class CategoryUpdate(BaseModel):
    new_name: str

    @field_validator('new_name')
    @classmethod
    def validate_new_name(cls, v):
        if not v or not v.strip():
            raise ValueError('분류명은 비어있을 수 없습니다.')
        if len(v.strip()) < 1:
            raise ValueError('분류명은 최소 1자 이상이어야 합니다.')
        if len(v.strip()) > 200:
            raise ValueError('분류명은 200자를 초과할 수 없습니다.')
        return v.strip()

router = APIRouter(
    prefix="/api/v1/classification",
    tags=["classification"],
    responses={404: {"description": "Not found"}},
)

@router.get("/list")
async def get_classifications(
    user: Dict[str, Any] = Depends(get_current_user)
):
    """분류 목록 조회 (wz_cate 테이블)"""
    try:
        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # wz_cate 테이블에서 분류 목록 조회
                cur.execute("""
                    SELECT
                        wzcateseq,
                        wzcatename,
                        wzorder
                    FROM wz_cate
                    WHERE wzvisible = 'Y' OR wzvisible IS NULL
                    ORDER BY wzorder, wzcateseq
                """)

                results = cur.fetchall()

                classifications = []
                for row in results:
                    cate_seq = row[0]

                    # 각 분류의 규정 수 계산 (현행 규정만)
                    cur.execute("""
                        SELECT COUNT(*)
                        FROM wz_rule
                        WHERE wzcateseq = %s
                        AND (wzNewFlag = '현행' OR wzNewFlag IS NULL)
                    """, (cate_seq,))

                    count = cur.fetchone()[0]

                    classifications.append({
                        "id": str(cate_seq),
                        "name": row[1].strip() if row[1] else '',
                        "order": row[2],
                        "count": count
                    })

                return {
                    "success": True,
                    "classifications": classifications,
                    "total": len(classifications)
                }

    except Exception as e:
        logger.error(f"Error getting classifications: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/create")
async def create_classification(
    category: CategoryCreate,
    user: Dict[str, Any] = Depends(get_current_user)
):
    """새 분류 생성"""
    try:
        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # 중복 확인
                cur.execute("""
                    SELECT COUNT(*) FROM wz_cate
                    WHERE wzcateseq = %s
                """, (category.chapter_number,))

                if cur.fetchone()[0] > 0:
                    return {
                        "success": False,
                        "error": "duplicate",
                        "message": f"제{category.chapter_number}장이 이미 존재합니다."
                    }

                # 새 분류 삽입
                cur.execute("""
                    INSERT INTO wz_cate (wzcateseq, wzcatename, wzparentseq, wzorder, wzvisible, wzcreatedby, wzmodifiedby)
                    VALUES (%s, %s, NULL, %s, 'Y', %s, %s)
                """, (category.chapter_number, category.name, category.chapter_number, user['username'], user['username']))

                conn.commit()

                return {
                    "success": True,
                    "message": f"제{category.chapter_number}장 '{category.name}'가 추가되었습니다."
                }

    except Exception as e:
        logger.error(f"Error creating classification: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/update/{cate_seq}")
async def update_classification(
    cate_seq: int,
    category: CategoryUpdate,
    user: Dict[str, Any] = Depends(get_current_user)
):
    """분류명 수정"""
    try:
        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # 분류 존재 확인
                cur.execute("""
                    SELECT wzcatename FROM wz_cate
                    WHERE wzcateseq = %s
                """, (cate_seq,))

                result = cur.fetchone()
                if not result:
                    raise HTTPException(status_code=404, detail="분류를 찾을 수 없습니다.")

                old_name = result[0].strip() if result[0] else ''

                # 분류명 업데이트
                cur.execute("""
                    UPDATE wz_cate
                    SET wzcatename = %s, wzmodifiedby = %s
                    WHERE wzcateseq = %s
                """, (category.new_name, user['username'], cate_seq))

                conn.commit()

                return {
                    "success": True,
                    "message": f"제{cate_seq}장이 '{old_name}'에서 '{category.new_name}'로 변경되었습니다."
                }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating classification: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/regulations/{cate_seq}")
async def get_regulations_by_category(
    cate_seq: int,
    user: Dict[str, Any] = Depends(get_current_user)
):
    """특정 분류의 규정 목록 조회"""
    try:
        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # 분류 존재 확인
                cur.execute("""
                    SELECT wzcatename FROM wz_cate
                    WHERE wzcateseq = %s
                """, (cate_seq,))

                cate_result = cur.fetchone()
                if not cate_result:
                    raise HTTPException(status_code=404, detail="분류를 찾을 수 없습니다.")

                cate_name = cate_result[0].strip() if cate_result[0] else ''

                # 해당 분류의 규정 목록 조회
                cur.execute("""
                    SELECT
                        wzruleseq,
                        wzname,
                        wzpubno,
                        wzFileJson,
                        wzNewFlag,
                        wzmgrdptorgcd,
                        wzmgrdptnm,
                        wzestabdate,
                        wzexecdate,
                        wzlastrevdate
                    FROM wz_rule
                    WHERE wzcateseq = %s
                    AND (wzNewFlag = '현행' OR wzNewFlag IS NULL)
                    ORDER BY wzpubno
                """, (cate_seq,))

                results = cur.fetchall()

                regulations = []
                for row in results:
                    regulations.append({
                        "id": row[0],
                        "name": row[1].strip() if row[1] else '',
                        "pubno": row[2].strip() if row[2] else '',
                        "filejson": row[3].strip() if row[3] else '',
                        "status": row[4] if row[4] else '현행',
                        "deptCode": row[5] if row[5] else '',
                        "deptName": row[6].strip() if row[6] else '',
                        "estabDate": row[7] if row[7] else '',
                        "execDate": row[8] if row[8] else '',
                        "lastRevDate": row[9] if row[9] else ''
                    })

                return {
                    "success": True,
                    "categoryId": cate_seq,
                    "categoryName": cate_name,
                    "regulations": regulations,
                    "total": len(regulations)
                }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting regulations by category: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/delete/{cate_seq}")
async def delete_classification(
    cate_seq: int,
    user: Dict[str, Any] = Depends(get_current_user)
):
    """분류 삭제"""
    try:
        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # 분류 존재 확인
                cur.execute("""
                    SELECT wzcatename FROM wz_cate
                    WHERE wzcateseq = %s
                """, (cate_seq,))

                result = cur.fetchone()
                if not result:
                    raise HTTPException(status_code=404, detail="분류를 찾을 수 없습니다.")

                cate_name = result[0].strip() if result[0] else ''

                # 해당 분류에 연결된 규정이 있는지 확인
                cur.execute("""
                    SELECT COUNT(*) FROM wz_rule
                    WHERE wzcateseq = %s
                """, (cate_seq,))

                regulation_count = cur.fetchone()[0]
                if regulation_count > 0:
                    return {
                        "success": False,
                        "error": "has_regulations",
                        "message": f"제{cate_seq}장 '{cate_name}'에 {regulation_count}개의 규정이 있어 삭제할 수 없습니다."
                    }

                # 분류 삭제
                cur.execute("""
                    DELETE FROM wz_cate
                    WHERE wzcateseq = %s
                """, (cate_seq,))

                conn.commit()

                return {
                    "success": True,
                    "message": f"제{cate_seq}장 '{cate_name}'가 삭제되었습니다."
                }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting classification: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/excel")
async def export_classification_excel(
    user: Dict[str, Any] = Depends(get_current_user)
):
    """
    사규 목차 Excel 다운로드 (별표 제1호 형식)
    DB의 wz_cate + wz_rule 데이터를 Excel로 내보내기
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                # 분류 목록 조회
                cur.execute("""
                    SELECT wzcateseq, wzcatename
                    FROM wz_cate
                    WHERE wzvisible = 'Y' OR wzvisible IS NULL
                    ORDER BY wzorder, wzcateseq
                """)
                categories = cur.fetchall()

                # 모든 현행 규정 조회
                cur.execute("""
                    SELECT
                        wzcateseq, wzpubno, wzname, wzmgrdptnm,
                        wzlastrevdate, wzestabdate
                    FROM wz_rule
                    WHERE wzNewFlag = '현행' OR wzNewFlag IS NULL
                    ORDER BY wzcateseq,
                        CAST(split_part(wzpubno, '-', 1) AS INTEGER),
                        CAST(split_part(wzpubno, '-', 2) AS INTEGER)
                """)
                regulations = cur.fetchall()

        # 규정을 분류별로 그룹화
        reg_by_cate = {}
        for reg in regulations:
            cate_seq = reg[0]
            if cate_seq not in reg_by_cate:
                reg_by_cate[cate_seq] = []
            reg_by_cate[cate_seq].append(reg)

        # Excel 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "사규 목차"

        # 스타일
        header_font = Font(name='KB금융체Text', size=14, bold=True)
        sub_font = Font(name='KB금융체Text', size=10)
        col_header_font = Font(name='KB금융체Text', size=10, bold=True)
        col_header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cat_font = Font(name='KB금융체Text', size=10, bold=True)
        body_font = Font(name='KB금융체Text', size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 제목
        ws.merge_cells('B1:H1')
        ws['B1'] = '(별표 제1호) 사규 목차'
        ws['B1'].font = header_font
        ws['B1'].alignment = Alignment(horizontal='center')

        today_str = datetime.now().strftime('%Y. %m. %d.')
        ws.merge_cells('B2:H2')
        ws['B2'] = f'({today_str} 기준)'
        ws['B2'].font = sub_font
        ws['B2'].alignment = Alignment(horizontal='center')

        # 헤더 행
        headers = ['번호', '구분', '인덱스', '사규명', '소관부서', '최근\n개정시행일자', '제정일자']
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 컬럼 폭 설정
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 14

        # 데이터 행
        row_num = 4
        seq_num = 1

        for cate_seq, cate_name in categories:
            regs = reg_by_cate.get(cate_seq, [])
            first_in_cat = True

            for reg in regs:
                pubno = (reg[1] or '').strip()
                name = (reg[2] or '').strip()
                dept = (reg[3] or '').strip()
                rev_date = reg[4]
                est_date = reg[5]

                # 번호
                ws.cell(row=row_num, column=2, value=seq_num).font = body_font
                ws.cell(row=row_num, column=2).border = thin_border
                ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')

                # 구분 (분류명) - 분류의 첫 규정에만 표시
                cat_cell = ws.cell(row=row_num, column=3)
                if first_in_cat:
                    cat_cell.value = f"{cate_seq}. {cate_name.strip()}"
                    cat_cell.font = cat_font
                    first_in_cat = False
                else:
                    cat_cell.value = None
                    cat_cell.font = body_font
                cat_cell.border = thin_border

                # 인덱스
                idx_cell = ws.cell(row=row_num, column=4, value=f"({pubno})" if pubno else '')
                idx_cell.font = body_font
                idx_cell.border = thin_border
                idx_cell.alignment = Alignment(horizontal='center')

                # 사규명
                name_cell = ws.cell(row=row_num, column=5, value=name)
                name_cell.font = body_font
                name_cell.border = thin_border

                # 소관부서
                dept_cell = ws.cell(row=row_num, column=6, value=dept)
                dept_cell.font = body_font
                dept_cell.border = thin_border
                dept_cell.alignment = Alignment(horizontal='center')

                # 개정일자
                rev_cell = ws.cell(row=row_num, column=7)
                if rev_date:
                    rev_cell.value = rev_date if isinstance(rev_date, datetime) else rev_date
                    rev_cell.number_format = 'YYYY-MM-DD'
                rev_cell.font = body_font
                rev_cell.border = thin_border
                rev_cell.alignment = Alignment(horizontal='center')

                # 제정일자
                est_cell = ws.cell(row=row_num, column=8)
                if est_date:
                    est_cell.value = est_date if isinstance(est_date, datetime) else est_date
                    est_cell.number_format = 'YYYY-MM-DD'
                est_cell.font = body_font
                est_cell.border = thin_border
                est_cell.alignment = Alignment(horizontal='center')

                row_num += 1
                seq_num += 1

        # 바이트 스트림으로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"KB_regulation_catalog_{datetime.now().strftime('%y%m%d')}.xlsx"
        filename_utf8 = f"KB신용정보_별표제1호_사규_목차_{datetime.now().strftime('%y%m%d')}.xlsx"
        from urllib.parse import quote
        encoded = quote(filename_utf8)

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"
            }
        )

    except Exception as e:
        logger.error(f"Error exporting classification excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import/excel")
async def import_classification_excel(
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(get_current_user)
):
    """
    사규 목차 Excel 업로드
    Excel에서 읽은 데이터로 wz_rule의 소관부서, 개정일, 제정일 등을 업데이트
    """
    try:
        import openpyxl

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Excel 파일(.xlsx)만 업로드 가능합니다.")

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        db_config = {
            'database': settings.DB_NAME,
            'user': settings.DB_USER,
            'password': settings.DB_PASSWORD,
            'host': settings.DB_HOST,
            'port': settings.DB_PORT
        }

        db_manager = DatabaseConnectionManager(**db_config)

        updated = 0
        errors = []

        with db_manager.get_connection() as conn:
            with conn.cursor() as cur:
                for row in ws.iter_rows(min_row=4, values_only=True):
                    # B=번호, C=구분, D=인덱스, E=사규명, F=소관부서, G=개정일, H=제정일
                    idx_raw = row[3] if len(row) > 3 else None  # D column (인덱스)
                    name = row[4] if len(row) > 4 else None     # E column (사규명)
                    dept = row[5] if len(row) > 5 else None     # F column (소관부서)
                    rev_date = row[6] if len(row) > 6 else None # G column
                    est_date = row[7] if len(row) > 7 else None # H column

                    if not idx_raw or not name:
                        continue

                    # 인덱스에서 공포번호 추출: (N-M) -> N-M
                    pubno = str(idx_raw).strip().strip('()')
                    if not pubno:
                        continue

                    try:
                        # wzpubno로 규정 찾기
                        cur.execute("""
                            SELECT wzruleseq FROM wz_rule
                            WHERE wzpubno = %s
                            AND (wzNewFlag = '현행' OR wzNewFlag IS NULL)
                        """, (pubno,))

                        result = cur.fetchone()
                        if not result:
                            errors.append(f"규정 '{pubno}' ({name}) 미발견")
                            continue

                        rule_seq = result[0]

                        # 업데이트 (사규명, 소관부서, 날짜 등)
                        update_fields = []
                        update_values = []

                        if name and str(name).strip():
                            update_fields.append("wzname = %s")
                            update_values.append(str(name).strip())

                        if dept and str(dept).strip():
                            update_fields.append("wzmgrdptnm = %s")
                            update_values.append(str(dept).strip())

                        if rev_date:
                            if isinstance(rev_date, datetime):
                                update_fields.append("wzlastrevdate = %s")
                                update_values.append(rev_date.strftime('%Y-%m-%d'))
                            elif isinstance(rev_date, str) and rev_date.strip():
                                update_fields.append("wzlastrevdate = %s")
                                update_values.append(rev_date.strip())

                        if est_date:
                            if isinstance(est_date, datetime):
                                update_fields.append("wzestabdate = %s")
                                update_values.append(est_date.strftime('%Y-%m-%d'))
                            elif isinstance(est_date, str) and est_date.strip():
                                update_fields.append("wzestabdate = %s")
                                update_values.append(est_date.strip())

                        if update_fields:
                            sql = f"UPDATE wz_rule SET {', '.join(update_fields)} WHERE wzruleseq = %s"
                            update_values.append(rule_seq)
                            cur.execute(sql, update_values)
                            updated += 1

                    except Exception as row_err:
                        errors.append(f"규정 '{pubno}' 처리 오류: {str(row_err)}")

                conn.commit()

        return {
            "success": True,
            "message": f"{updated}개 규정이 업데이트되었습니다.",
            "updated": updated,
            "errors": errors[:10] if errors else []
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing classification excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))